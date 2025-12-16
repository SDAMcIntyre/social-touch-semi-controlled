# src/pipeline_monitor_data_manager.py

import pandas as pd
from pathlib import Path
from typing import List, Optional

# Imports for Excel formatting
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle

class DataManager:
    """
    Handles in-memory state management and periodic flushing to XLSX.
    No longer handles locking; assumes a single owner (the Coordinator).
    """
    def __init__(self, report_file_path: Path):
        self.report_file_xlsx = report_file_path.with_suffix('.xlsx')
        self.report_file_csv = report_file_path.with_suffix('.csv') # For legacy migration
        self.report_file_xlsx.parent.mkdir(parents=True, exist_ok=True)

        self._df = pd.DataFrame() # The in-memory state

        self.STATUS_COLORS = {
            'SUCCESS': 'C6EFCE',  # Light Green
            'FAILURE': 'FFC7CE',  # Light Red
            'RUNNING': 'FFEB9C',  # Light Yellow
            'PENDING': 'D0D0D0',  # Gray
        }

    def initialize(self, columns: List[str]):
        """Initializes the in-memory dataframe and checks for existing files."""
        # 1. Try to load existing Excel
        if self.report_file_xlsx.exists():
            try:
                self._df = pd.read_excel(self.report_file_xlsx, engine='openpyxl')
                print(f"📊 Loaded existing report from: {self.report_file_xlsx}")
                return
            except Exception:
                print("⚠️ Existing Excel file corrupt or unreadable. Starting fresh.")

        # 2. Migration: Try to load legacy CSV
        if self.report_file_csv.exists():
            print(f"Legacy CSV report found. Migrating data...")
            self._df = pd.read_csv(self.report_file_csv)
            self.save_report() # Convert to xlsx immediately
            self.report_file_csv.unlink() # Delete legacy
            return

        # 3. New Report
        header = ['dataset'] + columns
        self._df = pd.DataFrame(columns=header)
        self.save_report()
        print(f"📊 Report initialized at: {self.report_file_xlsx}")

    def update_memory(self, dataset: str, event: str, status: str, message: str = "") -> pd.DataFrame:
        """
        Updates the in-memory DataFrame INSTANTLY. Does not write to disk.
        """
        value = f"{status}: {message}" if message else status
        
        if self._df.empty and 'dataset' not in self._df.columns:
             # Failsafe if init wasn't called, though Coordinator should handle this
             return self._df

        if event not in self._df.columns:
            # Dynamically add column if missing (optional safety)
            self._df[event] = None

        if dataset in self._df['dataset'].values:
            self._df.loc[self._df['dataset'] == dataset, event] = value
        else:
            new_row = {'dataset': dataset, event: value}
            new_row_df = pd.DataFrame([new_row])
            self._df = pd.concat([self._df, new_row_df], ignore_index=True)

        return self._df

    def get_dataframe(self) -> pd.DataFrame:
        return self._df

    def save_report(self):
        """
        Writes the current in-memory DataFrame to the formatted XLSX file.
        This is the expensive operation.
        """
        if self._df.empty:
            return

        try:
            with pd.ExcelWriter(self.report_file_xlsx, engine='openpyxl') as writer:
                self._df.to_excel(writer, index=False, sheet_name='Pipeline Status')
                worksheet = writer.sheets['Pipeline Status']

                # Auto-adjust column widths
                for idx, col in enumerate(self._df.columns, 1):
                    series = self._df[col]
                    if not series.empty:
                        # Calculate width based on max string length
                        max_len = max(
                            series.astype(str).map(len).max(),
                            len(str(series.name))
                        ) + 2
                        worksheet.column_dimensions[get_column_letter(idx)].width = max_len

                # Apply Conditional Formatting
                num_rows = len(self._df) + 1
                num_cols = len(self._df.columns)
                format_range = f"B2:{get_column_letter(num_cols)}{num_rows}"

                sorted_statuses = sorted(self.STATUS_COLORS.keys(), key=len, reverse=True)
                
                for status in sorted_statuses:
                    color_hex = self.STATUS_COLORS[status]
                    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                    dxf = DifferentialStyle(fill=fill)
                    formula = [f'SEARCH("{status}", B2)=1']
                    rule = FormulaRule(formula=formula, stopIfTrue=True)
                    rule.dxf = dxf
                    worksheet.conditional_formatting.add(format_range, rule)
        except PermissionError:
             print(f"⚠️ Could not write to {self.report_file_xlsx}. File might be open.")
        except Exception as e:
             print(f"🚨 CRITICAL: Failed to save report. Error: {e}")